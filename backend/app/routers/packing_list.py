"""
Router: Packing List OGL
Módulo de generación automática del Packing List para el cliente OGL.
Autor: AgroFlow Dev Team
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Request, Response
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, func
from typing import Optional, List, Dict, Any
from app.database import get_db
from app.utils.logging import logger
from app.models.pedido import PedidoComercial
from app.models.posicionamiento import Posicionamiento
from app.models.embarque import ControlEmbarque, ReporteEmbarques
from app.models.packing_list import EmisionPackingList, DetalleEmisionPackingList
from app.utils.formatters import get_peru_time
from app.dependencies.auth import OptionalUser, CurrentUser
from pydantic import BaseModel
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from copy import copy
import io
import os
import re
import difflib
from datetime import datetime

# Directorio donde se guardan los Excel generados para re-descarga
PL_STORAGE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "storage", "packing_lists")
os.makedirs(PL_STORAGE_DIR, exist_ok=True)

router = APIRouter(
    prefix="/api/v1/packing-list",
    tags=["Packing List OGL"]
)

# ---------------------------------------------------------------------------
# Mapeo de Recibidores Oficiales (Inge Daniel Rule 💎)
# ---------------------------------------------------------------------------
RECIPIENTS_DATA = {
    "ISS": {
        "full_name": "ISS (INTEGRATED SERVICE SOLUTIONS LTD)",
        "notify_id": "1038367"
    },
    "VDH": {
        "full_name": "VDH FRUITS BV",
        "notify_id": "1041374"
    },
    "VDH FRUITS BV": {
        "full_name": "VDH FRUITS BV",
        "notify_id": "1041374"
    }
}

# ---------------------------------------------------------------------------
# Constante: nombre del cliente objetivo
# ---------------------------------------------------------------------------
OGL_KEYWORD = "OGL"

# ---------------------------------------------------------------------------
# Helper: normalizar ORDEN_BETA para el JOIN con pedidos_comerciales
# ---------------------------------------------------------------------------
def strip_orden_beta(orden_beta: Optional[str]) -> Optional[str]:
    """Extrae la orden numérica pero evadiendo cruces con clientes que usan otros prefijos (ej. PE004)."""
    if not orden_beta:
        return None
    
    val_upper = orden_beta.strip().upper()
    
    # Si tiene letras, nos aseguramos de que sea solo BG, CO, BP, BAM, BU o BAA (prefijos de Beta)
    has_letters = any(c.isalpha() for c in val_upper)
    if has_letters:
        allowed_prefixes = ["BG", "CO", "BP", "BAM", "BU", "BAA"]
        if not any(prefix in val_upper for prefix in allowed_prefixes):
            # Si tiene letras y no es de Beta, devolvemos el valor original crudo.
            return val_upper
            
    # Si es puramente numérico o tiene prefijos autorizados, le quitamos las letras
    numeric = re.sub(r'[^0-9]', '', val_upper)
    return numeric if numeric else None

TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__),  # app/routers/
    "..", "..",                  # backend/
    "assets", "templates",
    "FORMATO PL - OGL.xlsx"
)

# ---------------------------------------------------------------------------
# Helper: Detectar si dos nombres de nave son la misma (Smart Match)
# ---------------------------------------------------------------------------
def is_same_ship(name1: str, name2: str, threshold: float = 0.75) -> bool:
    if not name1 or not name2: return False
    if name1 == "SIN NAVE" or name2 == "SIN NAVE": return False
    
    n1 = re.sub(r'[^A-Z0-9]', '', name1.upper())
    n2 = re.sub(r'[^A-Z0-9]', '', name2.upper())
    
    # 1. Identidad exacta normalizada
    if n1 == n2: return True
    
    # 2. Match por prefijo (Nombre del Barco)
    min_len = min(len(n1), len(n2))
    if min_len >= 10 and n1[:10] == n2[:10]:
        return True

    # 3. Similitud difusa
    return difflib.SequenceMatcher(None, n1, n2).ratio() >= threshold

# ---------------------------------------------------------------------------
# Schemas de respuesta
# ---------------------------------------------------------------------------
class NaveInfo(BaseModel):
    nave: str
    fuente: str      # "reporte_embarques" | "posicionamiento" | "consolidada"
    bookings: List[str]
    cultivos: List[str]

# ---------------------------------------------------------------------------
# GET /naves  ─  Lista todas las naves disponibles (Optimizado + Smart Match)
# ---------------------------------------------------------------------------
@router.get("/naves", response_model=List[NaveInfo])
def listar_naves_ogl(db: Session = Depends(get_db)):
    """
    Retorna naves únicas basándose en la unión de ReporteEmbarques y Posicionamiento.
    Optimizado para evitar el problema N+1 mediante JOINs estratégicos.
    """
    # Subquery para identificar bookings ya bloqueados por un PL activo
    blocked_sq = db.query(DetalleEmisionPackingList.booking).join(
        EmisionPackingList, DetalleEmisionPackingList.emision_id == EmisionPackingList.id
    ).filter(EmisionPackingList.estado == "ACTIVO").subquery()

    # Query Consolidada: Traemos Posicionamiento, su PedidoComercial (solo si es OGL)
    # y el ReporteEmbarques (nave_arribo) en una sola tanda.
    
    # Traer todos los posicionamientos que tienen booking
    posicionamientos = db.query(Posicionamiento).filter(
        Posicionamiento.BOOKING != None,
        ~Posicionamiento.BOOKING.in_(blocked_sq)
    ).all()

    # Traer todos los pedidos OGL para el cruce en memoria (mucho más rápido que N queries)
    pedidos_ogl = db.query(PedidoComercial).filter(
        PedidoComercial.cliente.ilike(f"%{OGL_KEYWORD}%")
    ).all()
    
    # Mapa de orden_beta_num -> pedido
    pedidos_map = {}
    for p in pedidos_ogl:
        ord_num = strip_orden_beta(p.orden_beta)
        if ord_num:
            pedidos_map[ord_num] = p

    # Traer todos los reportes para la nave final
    reportes = db.query(ReporteEmbarques.booking, ReporteEmbarques.nave_arribo).all()
    reporte_naves = {r.booking: r.nave_arribo for r in reportes if r.booking}

    nave_stats = {}

    for pos in posicionamientos:
        b = pos.BOOKING
        if not pos.ORDEN_BETA: continue
        
        ord_num = strip_orden_beta(pos.ORDEN_BETA)
        if not ord_num or ord_num not in pedidos_map: continue
        
        pedido_ogl = pedidos_map[ord_num]
        
        # Determinar Nave Final (Reporte > Posicionamiento)
        nave_final = reporte_naves.get(b).strip().upper() if reporte_naves.get(b) else None
        if not nave_final and pos.NAVE:
            nave_final = pos.NAVE.strip().upper()
            
        nave_final = nave_final if nave_final else "SIN NAVE"
        
        # Smart Match
        target_nave = nave_final
        if nave_final != "SIN NAVE":
            for existing_nave in nave_stats.keys():
                if existing_nave != "SIN NAVE" and is_same_ship(nave_final, existing_nave):
                    target_nave = existing_nave
                    break
        
        if target_nave not in nave_stats:
            nave_stats[target_nave] = {"bookings": [], "cultivos": set()}
        
        if b not in nave_stats[target_nave]["bookings"]:
            nave_stats[target_nave]["bookings"].append(b)
            # Determinar Cultivo
            c_final = pos.CULTIVO.strip().upper() if pos.CULTIVO else None
            if not c_final and pedido_ogl.cultivo:
                c_final = pedido_ogl.cultivo.strip().upper()
            if c_final:
                nave_stats[target_nave]["cultivos"].add(c_final)

    result = []
    for nave_name, stats in sorted(nave_stats.items()):
        if stats["bookings"]:
            result.append(NaveInfo(
                nave=nave_name,
                fuente="consolidada",
                bookings=stats["bookings"],
                cultivos=sorted(list(stats["cultivos"]))
            ))
    return result

@router.get("/bookings")
def listar_bookings_ogl(nave: str, db: Session = Depends(get_db)):
    """
    Lista todos los bookings cuya nave ACTUAL es la solicitada.
    Optimizado para evitar el problema N+1.
    """
    nave_upper = nave.strip().upper()
    
    # Subquery de bloqueos
    blocked_sq = db.query(DetalleEmisionPackingList.booking).join(
        EmisionPackingList, DetalleEmisionPackingList.emision_id == EmisionPackingList.id
    ).filter(EmisionPackingList.estado == "ACTIVO").subquery()

    # 1. Bookings confirmados en esta nave por el reporte
    bk_reporte = {r.booking for r in db.query(ReporteEmbarques.booking).filter(
        func.upper(func.trim(ReporteEmbarques.nave_arribo)) == nave_upper
    ).all() if r.booking}

    # 2. Bookings que dicen estar aquí en posicionamiento
    bk_pos_data = db.query(Posicionamiento).filter(
        func.upper(func.trim(Posicionamiento.NAVE)) == nave_upper
    ).all()
    bk_pos = {p.BOOKING for p in bk_pos_data if p.BOOKING}

    # 3. Fusión
    bookings_actuales = set(bk_reporte)
    for b in bk_pos:
        if b in bookings_actuales: continue
        movido = db.query(ReporteEmbarques).filter(
            ReporteEmbarques.booking == b,
            ReporteEmbarques.nave_arribo != None,
            func.upper(func.trim(ReporteEmbarques.nave_arribo)) != nave_upper
        ).first()
        if not movido:
            bookings_actuales.add(b)

    # 4. Enriquecimiento Optimizado
    pedidos_ogl = db.query(PedidoComercial).filter(PedidoComercial.cliente.ilike(f"%{OGL_KEYWORD}%")).all()
    pedidos_map = {strip_orden_beta(p.orden_beta): p for p in pedidos_ogl if strip_orden_beta(p.orden_beta)}
    
    pos_details = db.query(Posicionamiento).filter(
        Posicionamiento.BOOKING.in_(bookings_actuales),
        ~Posicionamiento.BOOKING.in_(blocked_sq.select())
    ).all()

    # Contenedores (ControlEmbarque)
    cont_data = db.query(ControlEmbarque.booking, ControlEmbarque.contenedor, ControlEmbarque.dam).filter(
        ControlEmbarque.booking.in_(bookings_actuales)
    ).all()
    cont_map = {c.booking: (c.contenedor, c.dam) for c in cont_data}

    resultado = []
    for pos in pos_details:
        ord_num = strip_orden_beta(pos.ORDEN_BETA)
        if not ord_num or ord_num not in pedidos_map: continue
        
        pedido = pedidos_map[ord_num]
        cont, dam = cont_map.get(pos.BOOKING, (None, None))

        resultado.append({
            "booking":       pos.BOOKING,
            "orden_beta":    pos.ORDEN_BETA,
            "contenedor":    cont,
            "dam":           dam,
            "port_id_orig":  pedido.port_id_orig,
            "port_id_dest":  pedido.port_id_dest,
            "variedad":      pedido.variedad,
            "total_cajas":   pedido.total_cajas,
            "presentacion":  pedido.presentacion,
            "pod":           pedido.pod,
            "consignatario": pedido.consignatario,
            "recibidor":     pedido.recibidor,
            "cliente":       pedido.cliente,
        })

    return sorted(resultado, key=lambda x: x["booking"])

# ---------------------------------------------------------------------------
# POST /generate/ogl  ─  Genera el Packing List CONSOLIDADO por NAVE
# ---------------------------------------------------------------------------
@router.post("/generate/ogl")
async def generate_packing_list_ogl(
    current_user: OptionalUser,
    nave: str = Form(...),
    confirmaciones: List[UploadFile] = File(...),
    termografos: Optional[UploadFile] = File(None),
    recibidor: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    try:
        nave_clean = nave.strip().upper()
        
        # 1. Obtener bookings (Reporte > Posicionamiento)
        bookings_reporte = db.query(ReporteEmbarques.booking).filter(
            func.upper(func.trim(ReporteEmbarques.nave_arribo)) == nave_clean
        ).all()
        bookings_reporte_set = {r.booking for r in bookings_reporte if r.booking}

        bookings_pos = db.query(Posicionamiento.BOOKING).filter(
            func.upper(func.trim(Posicionamiento.NAVE)) == nave_clean
        ).all()
        bookings_pos_set = {r.BOOKING for r in bookings_pos if r.BOOKING}

        bookings_set = set(bookings_reporte_set)
        for b in bookings_pos_set:
            if b in bookings_set: continue
            otra_nave = db.query(ReporteEmbarques).filter(
                ReporteEmbarques.booking == b,
                ReporteEmbarques.nave_arribo != None,
                func.upper(func.trim(ReporteEmbarques.nave_arribo)) != nave_clean
            ).first()
            if not otra_nave:
                bookings_set.add(b)

        if not bookings_set:
            raise HTTPException(status_code=404, detail=f"No se encontraron bookings para la nave '{nave}'")

        booking_data_map: Dict[str, Dict] = {}
        primer_pedido = None
        primer_pos = None

        for booking in sorted(bookings_set):
            pos = db.query(Posicionamiento).filter(Posicionamiento.BOOKING == booking).first()
            if not pos: continue

            pedido = None
            if pos.ORDEN_BETA:
                orden_numeric = strip_orden_beta(pos.ORDEN_BETA)
                if orden_numeric:
                    pedido = db.query(PedidoComercial).filter(
                        PedidoComercial.orden_beta == orden_numeric,
                        PedidoComercial.cliente.ilike(f"%{OGL_KEYWORD}%")
                    ).first()
            
            if not pedido: continue

            if recibidor and recibidor.strip():
                ped_recibidor = str(getattr(pedido, "recibidor", "") or "").strip().upper()
                if ped_recibidor != recibidor.strip().upper():
                    continue

            emb = db.query(ControlEmbarque).filter(ControlEmbarque.booking == booking).first()
            contenedor_fmt = format_container_ogl(emb.contenedor if emb else "")
            prog_date = getattr(pos, "FECHA_PROGRAMADA", None) or getattr(pos, "ETA", None) or datetime.now().date()

            booking_data_map[booking] = {
                "contenedor": contenedor_fmt, 
                "pedido": pedido, 
                "pos": pos,
                "fecha_prog": prog_date
            }
            if primer_pedido is None:
                primer_pedido = pedido; primer_pos = pos

        if not booking_data_map:
            raise HTTPException(status_code=404, detail="No se encontraron bookings OGL para consolidar")

        # Termógrafos
        termografos_map = {}
        if termografos:
            try:
                t_content = await termografos.read()
                t_df = pd.read_excel(io.BytesIO(t_content), engine="openpyxl", header=None)
                for _, t_row in t_df.iterrows():
                    if len(t_row) >= 14:
                        p_id_raw = str(t_row.iloc[13]).strip().upper() if pd.notna(t_row.iloc[13]) else ""
                        t_code_raw = str(t_row.iloc[12]).strip() if pd.notna(t_row.iloc[12]) else ""
                        if p_id_raw and t_code_raw and p_id_raw != "ID PALLET":
                            termografos_map[p_id_raw] = t_code_raw
            except Exception as e:
                logger.error(f"Error procesando termógrafos: {e}")

        # Cargar pallets
        agrupado_por_booking: Dict[str, List] = {b: [] for b in booking_data_map.keys()}
        contenedor_default = next(iter(booking_data_map.values()))["contenedor"]

        for conf_file in confirmaciones:
            content = await conf_file.read()
            try:
                df_raw = pd.read_excel(io.BytesIO(content), engine="openpyxl", header=None)
                header_index = 0
                for i in range(min(20, len(df_raw))):
                    row_values = [str(x).upper() for x in df_raw.iloc[i].values]
                    if any("PALLET" in val or "HU" in val for val in row_values):
                        header_index = i
                        break
                df = pd.read_excel(io.BytesIO(content), engine="openpyxl", header=header_index)
                df.columns = df.columns.str.strip().str.upper()
            except Exception as e:
                logger.error(f"Error al procesar archivo {conf_file.filename}: {e}")
                continue

            def find_col(df, kws):
                for c in [str(col) for col in df.columns]:
                    for kw in kws:
                        if kw.upper() in c.upper(): return c
                return None

            col_pallet  = find_col(df, ["PALLET", "HU", "ID PALLET"])
            col_booking = find_col(df, ["BOOKING", "DESPACHO"])
            col_orden_beta = find_col(df, ["ORDEN BETA", "ORDEN"])
            
            if not col_pallet:
                raise Exception(f"El archivo {conf_file.filename} no tiene columna de Pallets.")

            col_calibre = find_col(df, ["CALIBRE", "CALIDAD"])
            col_kilos   = find_col(df, ["KILOS", "PESO NETO", "NET"])
            col_cosecha = find_col(df, ["COSECHA", "HARVEST", "FECHA COSECHA"])
            col_proceso = find_col(df, ["PROCESO", "PROCESS", "FECHA PROCESO"])
            col_lote_ogl = find_col(df, ["LOTE CLIENTE (OGL)", "LOTE OGL", "CLIENT LOT"])
            col_cajas   = find_col(df, ["TOTAL DE CAJAS", "CAJAS", "BOXES", "QTY"])
            col_total_kilos = find_col(df, ["TOTAL KILOS", "NET WEIGHT", "PESO NETO TOTAL"])
            col_trazabilidad = find_col(df, ["CODIGO TRAZABILIDAD", "TRAZABILIDAD", "TRACEABILITY"])

            orden_to_bk = {}
            for bk, data in booking_data_map.items():
                num_obj = strip_orden_beta(data["pos"].ORDEN_BETA)
                if num_obj:
                    try:
                        n_str = str(int(num_obj))
                        orden_to_bk[n_str] = bk
                    except: pass

            last_valid_bk = None
            for _, row in df.iterrows():
                current_bk = ""
                if col_booking and pd.notna(row.get(col_booking)):
                    c_val = str(row.get(col_booking)).strip().upper()
                    if c_val in booking_data_map: current_bk = c_val
                if not current_bk and col_orden_beta and pd.notna(row.get(col_orden_beta)):
                    o_val = strip_orden_beta(str(row.get(col_orden_beta)))
                    if o_val:
                        try:
                            n_val = str(int(o_val))
                            if n_val in orden_to_bk: current_bk = orden_to_bk[n_val]
                        except: pass
                if current_bk: last_valid_bk = current_bk
                elif not current_bk and last_valid_bk: current_bk = last_valid_bk
                
                bk_f = current_bk
                if not bk_f or bk_f not in booking_data_map:
                    bk_f = next(iter(booking_data_map)) if len(booking_data_map) == 1 else "DESCONOCIDO"

                p_id = str(row.get(col_pallet)).strip() if pd.notna(row.get(col_pallet)) else ""
                if not p_id or p_id.lower() == "nan": continue

                if bk_f not in agrupado_por_booking: agrupado_por_booking[bk_f] = []

                def format_date_ogl(val):
                    if pd.isna(val) or str(val).strip() == "": return ""
                    try:
                        dt = pd.to_datetime(str(val).strip())
                        return dt.strftime("%d/%m/%Y")
                    except:
                        raw_str = str(val).strip()
                        if " " in raw_str: raw_str = raw_str.split(" ")[0]
                        if "-" in raw_str and len(raw_str.split("-")[0]) == 4:
                            parts = raw_str.split("-")
                            if len(parts) == 3: return f"{parts[2]}/{parts[1]}/{parts[0]}"
                        return raw_str

                agrupado_por_booking[bk_f].append({
                    "pallet": p_id,
                    "calibre": str(row.get(col_calibre, "")).strip() if col_calibre and pd.notna(row.get(col_calibre)) else "",
                    "kilos": row.get(col_kilos) if col_kilos else 0,
                    "total_kilos": row.get(col_total_kilos) if col_total_kilos else 0,
                    "cajas": row.get(col_cajas) if col_cajas else 0,
                    "cosecha": format_date_ogl(row.get(col_cosecha)),
                    "proceso": format_date_ogl(row.get(col_proceso)),
                    "lote_ogl": str(row.get(col_lote_ogl, "")).strip() if col_lote_ogl and pd.notna(row.get(col_lote_ogl)) else "",
                    "trazabilidad": str(row.get(col_trazabilidad, "")).strip() if col_trazabilidad and pd.notna(row.get(col_trazabilidad)) else "",
                })

        # 3. Escribir Excel
        lista_ordenada = sorted(booking_data_map.items(), key=lambda x: x[1]["fecha_prog"])
        template_path = os.path.normpath(TEMPLATE_PATH)
        wb = openpyxl.load_workbook(template_path, keep_vba=False)
        ws = wb.active

        def safe_write(ws, cell_ref, value):
            try:
                cell = ws[cell_ref]
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    cell.value = value
            except: pass

        ahora = get_peru_time()
        
        # WK ID logic
        pl_id = f"WK{ahora.isocalendar()[1]}1"
        if primer_pos and primer_pos.ETA:
            semana_eta = primer_pos.ETA.isocalendar()[1]
            anio_eta = primer_pos.ETA.year
            pos_semana = db.query(Posicionamiento).filter(
                func.extract('week', Posicionamiento.ETA) == semana_eta,
                func.extract('year', Posicionamiento.ETA) == anio_eta
            ).all()
            nave_etas = {}
            for p in pos_semana:
                if not p.NAVE or not p.ETA or not p.ORDEN_BETA: continue
                num_ord = strip_orden_beta(p.ORDEN_BETA)
                if num_ord:
                    pedido_ogl = db.query(PedidoComercial).filter(
                        PedidoComercial.orden_beta == num_ord,
                        PedidoComercial.cliente.ilike(f"%{OGL_KEYWORD}%")
                    ).first()
                    if pedido_ogl:
                        rep = db.query(ReporteEmbarques).filter(ReporteEmbarques.booking == p.BOOKING).first()
                        n_name = (rep.nave_arribo if rep and rep.nave_arribo else p.NAVE).strip().upper()
                        etd_val = p.ETD if p.ETD else p.ETA
                        if n_name not in nave_etas or etd_val < nave_etas[n_name]:
                            nave_etas[n_name] = etd_val
            if nave_clean not in nave_etas:
                nave_etas[nave_clean] = primer_pos.ETD if primer_pos else (ahora.date())
            naves_ordenadas = sorted(nave_etas.items(), key=lambda x: (x[1], x[0]))
            correlativo = 1
            for i, (n_name, _) in enumerate(naves_ordenadas):
                if n_name == nave_clean:
                    correlativo = i + 1
                    break
            pl_id = f"WK{str(semana_eta).zfill(2)}{correlativo}"

        # Cabecera
        ws['C2'].value = "1101613"
        ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
        safe_write(ws, "C3", "COMPLEJO AGROINDUSTRIAL BETA S.A.")
        safe_write(ws, "C4", pl_id)
        safe_write(ws, "C5", ahora.strftime("%d/%m/%Y"))
        safe_write(ws, "C6", "CIF")
        safe_write(ws, "C7", "VESSEL")
        nave_final = primer_pos.NAVE if primer_pos else nave_clean
        reporte_n = db.query(ReporteEmbarques).filter(ReporteEmbarques.booking == next(iter(bookings_set))).first()
        if reporte_n and reporte_n.nave_arribo: nave_final = reporte_n.nave_arribo
        safe_write(ws, "C8", nave_final)

        if primer_pedido:
            recibidor_raw = (primer_pedido.recibidor or "").strip().upper()
            recipient_info = RECIPIENTS_DATA.get(recibidor_raw)
            if not recipient_info:
                if "VDH" in recibidor_raw: recipient_info = RECIPIENTS_DATA.get("VDH")
                elif "ISS" in recibidor_raw: recipient_info = RECIPIENTS_DATA.get("ISS")
            if recipient_info:
                safe_write(ws, "C10", recipient_info["notify_id"])
                safe_write(ws, "C11", recipient_info["full_name"])
            else:
                safe_write(ws, "C11", primer_pedido.recibidor or "")
            safe_write(ws, "C12", primer_pedido.port_id_orig or "")
            safe_write(ws, "C14", primer_pedido.port_id_dest or "")
            safe_write(ws, "C15", primer_pedido.pod or "")
        
        if primer_pos:
            safe_write(ws, "C13", primer_pos.POL or "")
            safe_write(ws, "C16", primer_pos.ETD.strftime("%d/%m/%Y") if primer_pos.ETD else "")
            safe_write(ws, "C17", primer_pos.ETA.strftime("%d/%m/%Y") if primer_pos.ETA else "")

        def safe_float(val):
            try: return float(val) if pd.notna(val) else 0.0
            except: return 0.0

        GRID_START_ROW = 21
        fila_secuencial = 1
        therms_written = set()
        total_cajas_acum = 0
        total_gross_acum = 0.0
        tipo_emision = "OTRO"
        
        for bk_id, _ in lista_ordenada:
            pedido = booking_data_map[bk_id]["pedido"]
            pos_bk = db.query(Posicionamiento).filter(Posicionamiento.BOOKING == bk_id).first()
            planta_llenado_raw = pos_bk.PLANTA_LLENADO.strip() if pos_bk and pos_bk.PLANTA_LLENADO else ""
            planta_llenado_up = planta_llenado_raw.upper()

            for item in agrupado_por_booking.get(bk_id, []):
                fila_e = GRID_START_ROW + (fila_secuencial - 1)
                ws.cell(row=fila_e, column=3).value = item["pallet"]
                ws.cell(row=fila_e, column=4).value = booking_data_map[bk_id]["contenedor"]
                ws.cell(row=fila_e, column=6).value = pedido.product.strip() if pedido and pedido.product else ""
                ws.cell(row=fila_e, column=7).value = pedido.variedad.strip() if pedido and pedido.variedad else ""
                ws.cell(row=fila_e, column=8).value = item["calibre"]
                peso_val = pedido.peso_por_caja if pedido else ""
                if peso_val is not None and str(peso_val).strip() != "":
                    ws.cell(row=fila_e, column=9).value = f"{str(peso_val).strip()} KG"
                p_id_match = str(item["pallet"]).strip().upper()
                if p_id_match in termografos_map and p_id_match not in therms_written:
                    ws.cell(row=fila_e, column=13).value = termografos_map[p_id_match]
                    therms_written.add(p_id_match)
                cajas_num = safe_float(item["cajas"])
                factor = 4.2
                cultivo_up = (pedido.cultivo or "").upper() if pedido else ""
                peso_nominal = safe_float(pedido.peso_por_caja) if pedido else 0.0
                if "PALTA" in cultivo_up or "AVOCADO" in cultivo_up:
                    ws.cell(row=fila_e, column=10).value = "LOOSE"
                    if peso_nominal == 10: 
                        factor = 10.97
                        tipo_emision = "PALTA_10KG"
                    elif peso_nominal == 4: 
                        factor = 4.3
                        tipo_emision = "PALTA_4KG"
                gross_val = round(cajas_num * factor, 2)
                ws.cell(row=fila_e, column=14).value = gross_val
                total_cajas_acum += cajas_num
                total_gross_acum += gross_val
                ws.cell(row=fila_e, column=15).value = item["total_kilos"]
                ws.cell(row=fila_e, column=16).value = item["cosecha"]
                ws.cell(row=fila_e, column=17).value = item["proceso"]
                ws.cell(row=fila_e, column=18).value = item["lote_ogl"]
                ws.cell(row=fila_e, column=19).value = "Complejo Agroindustrial"
                if "ICA" in planta_llenado_up: ws.cell(row=fila_e, column=20).value = "7751043044355"
                ws.cell(row=fila_e, column=21).value = pedido.caja_por_pallet if pedido else ""
                ws.cell(row=fila_e, column=22).value = item["cajas"]
                ws.cell(row=fila_e, column=23).value = "COMPLEJO AGROINDUSTRIAL BETA S.A."
                ws.cell(row=fila_e, column=24).value = "4050373153151"
                ws.cell(row=fila_e, column=25).value = planta_llenado_raw
                ws.cell(row=fila_e, column=26).value = item["trazabilidad"]
                fila_secuencial += 1

        if agrupado_por_booking.get("DESCONOCIDO"):
            planta_llenado_raw = primer_pos.PLANTA_LLENADO.strip() if primer_pos and primer_pos.PLANTA_LLENADO else ""
            planta_llenado_up = planta_llenado_raw.upper()
            for item in agrupado_por_booking["DESCONOCIDO"]:
                fila_e = GRID_START_ROW + (fila_secuencial - 1)
                ws.cell(row=fila_e, column=3).value = item["pallet"]; ws.cell(row=fila_e, column=4).value = contenedor_default
                ws.cell(row=fila_e, column=6).value = primer_pedido.product if primer_pedido else ""; ws.cell(row=fila_e, column=7).value = primer_pedido.variedad if primer_pedido else ""
                ws.cell(row=fila_e, column=8).value = item["calibre"]
                peso_val = primer_pedido.peso_por_caja if primer_pedido else ""
                if peso_val is not None and str(peso_val).strip() != "": ws.cell(row=fila_e, column=9).value = f"{str(peso_val).strip()} KG"
                if primer_pedido and ("PALTA" in (primer_pedido.cultivo or "").upper() or "AVOCADO" in (primer_pedido.cultivo or "").upper()):
                    ws.cell(row=fila_e, column=10).value = "LOOSE"
                p_id_match = str(item["pallet"]).strip().upper()
                if p_id_match in termografos_map and p_id_match not in therms_written:
                    ws.cell(row=fila_e, column=13).value = termografos_map[p_id_match]
                    therms_written.add(p_id_match)
                ws.cell(row=fila_e, column=14).value = round(safe_float(item["cajas"]) * 4.2, 2)
                ws.cell(row=fila_e, column=15).value = item["total_kilos"]; ws.cell(row=fila_e, column=16).value = item["cosecha"]; ws.cell(row=fila_e, column=17).value = item["proceso"]
                ws.cell(row=fila_e, column=18).value = item["lote_ogl"]; ws.cell(row=fila_e, column=19).value = "Complejo Agroindustrial"
                if "ICA" in planta_llenado_up: ws.cell(row=fila_e, column=20).value = "7751043044355"
                ws.cell(row=fila_e, column=21).value = primer_pedido.caja_por_pallet if primer_pedido else ""; ws.cell(row=fila_e, column=22).value = item["cajas"]
                ws.cell(row=fila_e, column=23).value = "COMPLEJO AGROINDUSTRIAL BETA S.A."; ws.cell(row=fila_e, column=24).value = "4050373153151"
                ws.cell(row=fila_e, column=25).value = planta_llenado_raw; ws.cell(row=fila_e, column=26).value = item["trazabilidad"]
                fila_secuencial += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Validaciones Finales
        warnings = []
        if tipo_emision == "PALTA_10KG":
            if int(total_cajas_acum) != 2400: warnings.append(f"ALERTA: Cajas ({int(total_cajas_acum)}) != 2400.")
            if round(total_gross_acum, 2) != 26328.00: warnings.append(f"ALERTA: Peso Bruto ({round(total_gross_acum, 2)}) != 26328.")
        elif tipo_emision == "PALTA_4KG":
            if int(total_cajas_acum) != 5280: warnings.append(f"ALERTA: Cajas ({int(total_cajas_acum)}) != 5280.")
            if round(total_gross_acum, 2) != 22704.00: warnings.append(f"ALERTA: Peso Bruto ({round(total_gross_acum, 2)}) != 22704.")
        
        safe_nave = re.sub(r'[\\/*?:"<>|]', "_", nave_clean).replace(' ', '_')
        filename = f"Packing List_OGL_MAESTRO_{safe_nave}_{pl_id}.xlsx"
        file_bytes = output.getvalue()
        storage_path = os.path.join(PL_STORAGE_DIR, filename)
        with open(storage_path, "wb") as f_disk: f_disk.write(file_bytes)

        # Auditoría
        try:
            nombre_usuario = current_user.usuario.upper() if current_user else "SISTEMA"
            nueva_emision = EmisionPackingList(usuario=nombre_usuario, nave=nave_clean, estado="ACTIVO", archivo_nombre=filename)
            db.add(nueva_emision); db.flush()
            for bk_id in bookings_set:
                db.add(DetalleEmisionPackingList(emision_id=nueva_emision.id, booking=bk_id))
            db.commit()
        except Exception as inner_e:
            db.rollback(); logger.error(f"Error auditoría: {inner_e}")
            raise HTTPException(status_code=500, detail="Error al guardar historial")

        headers = {"Content-Disposition": f"attachment; filename={filename}", "Access-Control-Expose-Headers": "X-PL-Warnings"}
        if warnings:
            import json
            headers["X-PL-Warnings"] = json.dumps(warnings)
        return Response(content=file_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    except HTTPException: raise
    except Exception as e: 
        import traceback
        logger.error(f"FATAL ERROR en generar_pl_ogl: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {str(e)}")

def format_container_ogl(code: str) -> str:
    if not code or len(code) < 5: return code or ""
    code = code.strip().upper()
    if " " in code: return code
    return f"{code[:4]} {code[4:]}"

@router.get("/historial")
def obtener_historial_pl(db: Session = Depends(get_db)):
    emisiones = db.query(EmisionPackingList).order_by(EmisionPackingList.fecha_generacion.desc()).limit(100).all()
    resultado = []
    for em in emisiones:
        archivo_disponible = os.path.exists(os.path.join(PL_STORAGE_DIR, em.archivo_nombre)) if em.archivo_nombre else False
        bk_list = [d.booking for d in em.detalles]
        pos_data = db.query(Posicionamiento.ORDEN_BETA).filter(Posicionamiento.BOOKING.in_(bk_list)).all()
        ord_list = sorted(list(set([p[0] for p in pos_data if p[0]])))
        resultado.append({
            "id": em.id, "fecha": em.fecha_generacion, "usuario": em.usuario or "Sistema", "nave": em.nave,
            "estado": em.estado, "archivo": em.archivo_nombre, "archivo_disponible": archivo_disponible,
            "motivo_anulacion": em.motivo_anulacion, "usuario_anulacion": em.usuario_anulacion,
            "bookings": bk_list, "ordenes": ord_list
        })
    return {"items": resultado}

class AnularPLRequest(BaseModel):
    motivo: str

@router.patch("/{id}/anular")
def anular_pl(id: int, req: AnularPLRequest, current_user: CurrentUser, db: Session = Depends(get_db)):
    emision = db.query(EmisionPackingList).filter(EmisionPackingList.id == id).first()
    if not emision: raise HTTPException(status_code=404, detail="Packing List no encontrado")
    if emision.estado == "ANULADO": raise HTTPException(status_code=400, detail="Ya se encuentra anulado")
    emision.estado = "ANULADO"; emision.motivo_anulacion = req.motivo
    emision.usuario_anulacion = current_user.usuario.upper() if current_user else "SISTEMA"
    try: db.commit()
    except: db.rollback(); raise HTTPException(status_code=500, detail="Error al anular")
    return {"message": "Anulado correctamente", "id": id}

@router.get("/{id}/descargar")
def descargar_pl(id: int, db: Session = Depends(get_db)):
    emision = db.query(EmisionPackingList).filter(EmisionPackingList.id == id).first()
    if not emision or not emision.archivo_nombre: raise HTTPException(status_code=404, detail="Archivo no encontrado")
    path_disco = os.path.join(PL_STORAGE_DIR, emision.archivo_nombre)
    if not os.path.exists(path_disco): raise HTTPException(status_code=404, detail="Archivo no disponible en disco")
    return FileResponse(path=path_disco, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=emision.archivo_nombre, headers={"Access-Control-Expose-Headers": "Content-Disposition"})
