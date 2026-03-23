import os
import openpyxl
import re
import io
from datetime import datetime, date
from typing import List, Optional
from collections import defaultdict
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.database import get_db
from app.models.ref_posicionamiento import RefPosicionamiento
from app.models.ie_models import CatClienteIE, RegistroIE
from app.models.ref_booking_dam import RefBookingDam
from app.models.packing_ogl import CuadroPedido, PackingConfirmacion, PackingTermografo

router = APIRouter(prefix="/api/v1/packing-ogl", tags=["Packing OGL"])

# --- HELPERS ---
def clean_numeric(val) -> Optional[int]:
    try:
        nums = re.sub(r'\D', '', str(val))
        return int(nums) if nums else None
    except: return None

def format_order(num: int) -> str:
    return f"BAM-{str(num).zfill(3)}"

# --- ENDPOINTS ---

@router.get("/naves")
def get_ogl_naves(db: Session = Depends(get_db)):
    """Lista naves únicas que tienen órdenes de clientes OGL"""
    # Filtramos clientes OGL en CatClienteIE
    ogl_clients = db.query(CatClienteIE).filter(CatClienteIE.nombre_comercial.ilike("%OGL%")).all()
    if not ogl_clients:
        return []
    
    ogl_names = [c.nombre_comercial for c in ogl_clients]
    
    # Buscamos naves en posicionamiento para esos clientes
    # El usuario quiere ver las naves así estén finalizadas para poder ver el historial o descargar de nuevo
    naves = db.query(RefPosicionamiento.nave).filter(
        RefPosicionamiento.cliente.in_(ogl_names),
        RefPosicionamiento.nave.isnot(None),
        RefPosicionamiento.nave != ""
    ).distinct().all()
    
    return [n[0] for n in naves if n[0]]

@router.get("/orders")
def get_ogl_orders(nave: str, db: Session = Depends(get_db)):
    """Lista las órdenes OGL (pendientes y finalizadas) para una nave específica"""
    ogl_clients = db.query(CatClienteIE).filter(CatClienteIE.nombre_comercial.ilike("%OGL%")).all()
    ogl_names = [c.nombre_comercial for c in ogl_clients]
    
    rows = db.query(RefPosicionamiento).filter(
        RefPosicionamiento.nave == nave,
        RefPosicionamiento.cliente.in_(ogl_names)
    ).all()
    
    return [
        {
            "orden": r.orden_beta_final or format_order(clean_numeric(r.o_beta_inicial)),
            "booking": r.booking,
            "finalizado": r.packing_ogl_finalizado
        }
        for r in rows
    ]

@router.post("/upload-confirmacion")
async def upload_confirmacion(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Parsea el Excel de Confirmación y puebla la DB. Bloquea si la orden ya está finalizada."""
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    if "FORMATO" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="No se encontró la hoja 'FORMATO'")
    
    sheet = wb["FORMATO"]
    header_row = 6
    
    processed_orders = set()
    rows_added = 0
    
    for r in range(header_row + 1, sheet.max_row + 1):
        order_raw = sheet.cell(row=r, column=5).value
        pallet_id = str(sheet.cell(row=r, column=6).value or "").strip()
        
        if not order_raw or not pallet_id: continue
        
        num_order = clean_numeric(order_raw)
        if not num_order: continue
        
        # User wants to preserve his exact format (BG-004, etc.)
        order_beta = str(order_raw).strip().upper()

        # VALIDACIÓN: ¿Está ya finalizada? (Usando el número puramente para el cruce)
        pos = db.query(RefPosicionamiento).filter(
            (RefPosicionamiento.o_beta_inicial.ilike(f"%{num_order}%")) |
            (RefPosicionamiento.orden_beta_final.ilike(f"%{num_order}%"))
        ).filter(RefPosicionamiento.packing_ogl_finalizado == True).first()
        
        if pos:
            raise HTTPException(
                status_code=403, 
                detail=f"La orden {order_beta} ya fue finalizada en un Packing List previo."
            )

        processed_orders.add(order_beta)
        
        # Upsert Pallet
        conf = db.query(PackingConfirmacion).filter(
            PackingConfirmacion.orden_beta == order_beta,
            PackingConfirmacion.pallet_id == pallet_id
        ).first()
        
        if not conf:
            conf = PackingConfirmacion(orden_beta=order_beta, pallet_id=pallet_id)
            db.add(conf)
            
        conf.fecha_cosecha = sheet.cell(row=r, column=13).value
        conf.fecha_proceso = sheet.cell(row=r, column=14).value
        conf.lote_ogl = str(sheet.cell(row=r, column=15).value or "").strip()
        conf.codigo_trazabilidad = str(sheet.cell(row=r, column=16).value or "").strip()
        conf.calibre = str(sheet.cell(row=r, column=23).value or "").strip()
        conf.archivo_nombre = file.filename
        
        try: conf.total_cajas = int(sheet.cell(row=r, column=33).value or 0)
        except: pass
        try: conf.peso_neto = float(sheet.cell(row=r, column=34).value or 0)
        except: pass
        
        rows_added += 1
        
    db.commit()
    return {"ok": True, "orders": list(processed_orders), "pallets_added": rows_added}

@router.post("/upload-termografos")
async def upload_termografos(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Puebla tabla de termógrafos"""
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet = wb.active # Hoja MARITIMO por defecto
    
    header_row = 2
    added = 0
    for r in range(header_row + 1, sheet.max_row + 1):
        order_raw = sheet.cell(row=r, column=6).value
        term_code = str(sheet.cell(row=r, column=13).value or "").strip()
        pallet_id = str(sheet.cell(row=r, column=14).value or "").strip()
        
        if not order_raw or not term_code or term_code == "None": continue
        
        num_order = clean_numeric(order_raw)
        if not num_order: continue
        
        # Preservar el nombre tal cual viene en el excel
        order_beta = str(order_raw).strip().upper()
        
        # Upsert Termografo
        term = db.query(PackingTermografo).filter(
            PackingTermografo.orden_beta == order_beta,
            PackingTermografo.pallet_id == pallet_id
        ).first()
        
        if not term:
            term = PackingTermografo(orden_beta=order_beta, pallet_id=pallet_id)
            db.add(term)
        
        term.codigo_termografo = term_code
        added += 1
        
    db.commit()
    return {"ok": True, "added": added}

@router.get("/generate/{nave}")
def generate_packing_ogl(nave: str, db: Session = Depends(get_db)):
    """Genera el Excel final consolidado por NAVE"""
    # 1. Buscar órdenes asociadas a la nave (Solo OGL)
    ogl_clients = db.query(CatClienteIE).filter(CatClienteIE.nombre_comercial.ilike("%OGL%")).all()
    ogl_names = [c.nombre_comercial for c in ogl_clients]
    
    orders = db.query(RefPosicionamiento).filter(
        RefPosicionamiento.nave == nave,
        RefPosicionamiento.cliente.in_(ogl_names)
    ).all()
    
    if not orders:
        raise HTTPException(status_code=404, detail=f"No hay órdenes OGL para la nave {nave}")
    
    # 2. Cargar Plantilla
    current_dir = os.path.dirname(os.path.abspath(__file__))
    # El archivo está en backend/app/routers/packing_ogl.py
    # Los assets están en backend/assets/templates/
    template_path = os.path.join(current_dir, "..", "..", "assets", "templates", "FORMATO PL - OGL.xlsx")
    
    if not os.path.exists(template_path):
        # Fallback por si la estructura cambia o estamos en local dev root
        template_path = os.path.join(os.getcwd(), "backend", "assets", "templates", "FORMATO PL - OGL.xlsx")
        
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail=f"No se encontró la plantilla en {template_path}")

    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    # 3. Rellenar Cabecera (Basado en la primera orden o promedio)
    main_order = orders[0]
    
    # B11: Notify First Line
    notify_full = ""
    cliente_ie = db.query(CatClienteIE).filter(CatClienteIE.nombre_comercial == main_order.cliente).first()
    if cliente_ie and cliente_ie.notificante_bl:
        notify_full = cliente_ie.notificante_bl.split("\n")[0]
    
    sheet["B2"] = "1101613" # Supplier ID
    sheet["B3"] = "COMPLEJO AGROINDUSTRIAL BETA S.A."
    sheet["B4"] = "WK471" # Default según user
    sheet["B5"] = datetime.now().strftime("%d/%m/%Y")
    sheet["B6"] = main_order.incoterm
    sheet["B7"] = "VESSEL"
    sheet["B8"] = nave
    sheet["B10"] = "1041374"
    sheet["B11"] = notify_full
    sheet["B12"] = "PEPAI"
    sheet["B13"] = main_order.pol
    sheet["B14"] = "NLRTM"
    sheet["B15"] = main_order.destino_booking
    sheet["B16"] = main_order.etd_booking
    sheet["B17"] = main_order.eta_booking
    
    # 4. Normalización y Pre-carga (Cruce Robusto)
    # Ordenamos órdenes numéricamente para que salgan en orden (BAM-01, BAM-02...)
    orders.sort(key=lambda x: clean_numeric(x.orden_beta_final or x.o_beta_inicial) or 0)

    all_normalized_nums = []
    for posic in orders:
        posic.packing_ogl_finalizado = True
        num = clean_numeric(posic.orden_beta_final or posic.o_beta_inicial)
        if num: all_normalized_nums.append(num)
    db.commit()

    # Buscamos en las tablas usando el formato BAM-XXX (zfill 3) que es el estándar de guardado
    target_keys = [format_order(n) for n in all_normalized_nums]

    cuadros_map = {c.orden_beta: c for c in db.query(CuadroPedido).filter(CuadroPedido.orden_beta.in_(target_keys)).all()}
    
    # Cruce por número normalizado (int) para evitar fallos de ceros a la izquierda
    confs_all = db.query(PackingConfirmacion).filter(PackingConfirmacion.orden_beta.in_(target_keys)).all()
    confs_map = defaultdict(list)
    for c in confs_all:
        c_num = clean_numeric(c.orden_beta)
        if c_num: confs_map[c_num].append(c)
        
    all_bookings = [o.booking for o in orders if o.booking]
    dams_map = {d.booking: d for d in db.query(RefBookingDam).filter(RefBookingDam.booking.in_(all_bookings)).all()}
    
    terms_all = db.query(PackingTermografo).filter(PackingTermografo.orden_beta.in_(target_keys)).all()
    terms_map = defaultdict(list)
    for t in terms_all:
        t_num = clean_numeric(t.orden_beta)
        if t_num: terms_map[t_num].append(t)

    # 5. Detalle de Pallets (Iterativo sobre data pre-cargada)
    current_row = 21
    total_pallets_all = 0
    
    for posic in orders:
        num_key = clean_numeric(posic.orden_beta_final or posic.o_beta_inicial)
        if not num_key: continue
        
        bam_key = format_order(num_key)
        cp = cuadros_map.get(bam_key)
        confirmaciones = confs_map.get(num_key, [])
        dam = dams_map.get(posic.booking)
        
        cont_no = f"MMAU {dam.dam}" if dam and dam.dam else (posic.nro_fcl or "")
        
        # --- CÁLCULO DE PESOS ---
        total_u = 0
        try: total_u = int(re.sub(r'[^0-9]', '', str(posic.total_unidades or "0")))
        except: pass
        
        # Peso Caja: Prioridad CuadroPedidos -> cj_kg -> Default 3.8
        peso_caja_val = 3.8
        if cp and cp.peso_caja:
            peso_caja_val = float(cp.peso_caja)
        elif posic.cj_kg:
            m_p = re.search(r'([0-9.]+)', str(posic.cj_kg))
            if m_p: peso_caja_val = float(m_p.group(1))

        # Cálculo Total
        p_neto_total = float(posic.peso_neto or (total_u * peso_caja_val))
        p_bruto_total = float(posic.peso_bruto or (p_neto_total + (total_u * 0.4)))
        
        total_p_en_orden = float(posic.total_pallet or len(confirmaciones) or 1)
        peso_bruto_pallet = (p_bruto_total / total_p_en_orden) if total_p_en_orden > 0 else 0
        
        # Mapa de termógrafos de ESTA orden para búsqueda rápida
        terms_de_orden = {str(t.pallet_id).replace('AR-', ''): t for t in terms_map.get(num_key, [])}
        
        for conf in confirmaciones:
            # Buscar Termógrafo usando el mapa pre-cargado
            conf_clean_id = str(conf.pallet_id).replace('GW-', '')
            term = terms_de_orden.get(conf_clean_id)
            
            term_txt = f"THERMOGRAPH: {term.codigo_termografo}" if term else ""
            base_notes = cp.additional_info if cp and cp.additional_info else "WITHOUT LABEL"
            final_note = f"{base_notes} - {term_txt}" if term_txt else base_notes

            # Llenar fila
            sheet.cell(row=current_row, column=2).value = "" # Order Ref (User said empty)
            sheet.cell(row=current_row, column=3).value = conf.pallet_id
            sheet.cell(row=current_row, column=4).value = cont_no
            sheet.cell(row=current_row, column=5).value = "" # Article Ref
            sheet.cell(row=current_row, column=6).value = cp.product if cp else "POMEGRANATE"
            sheet.cell(row=current_row, column=7).value = posic.variedad
            sheet.cell(row=current_row, column=8).value = conf.calibre
            sheet.cell(row=current_row, column=9).value = cp.peso_caja if cp else "3.8 KG"
            sheet.cell(row=current_row, column=10).value = "" # Pack Type
            sheet.cell(row=current_row, column=11).value = "" # Brand
            sheet.cell(row=current_row, column=12).value = "" # Harvest Product
            sheet.cell(row=current_row, column=13).value = final_note
            sheet.cell(row=current_row, column=14).value = round(peso_bruto_pallet, 3)
            
            # Columnas adicionales (Net Weight, Dates, etc.)
            # Basado en el mapeo extendido del user
            sheet.cell(row=current_row, column=15).value = conf.peso_neto
            sheet.cell(row=current_row, column=16).value = conf.fecha_cosecha
            sheet.cell(row=current_row, column=17).value = conf.fecha_proceso
            sheet.cell(row=current_row, column=18).value = conf.lote_ogl
            sheet.cell(row=current_row, column=19).value = "Complejo Agroindustrial"
            sheet.cell(row=current_row, column=20).value = cp.house_gln if cp else ""
            sheet.cell(row=current_row, column=21).value = cp.cajas_por_pallet if cp else ""
            sheet.cell(row=current_row, column=22).value = conf.total_cajas
            sheet.cell(row=current_row, column=23).value = "COMPLEJO AGROINDUSTRIAL BETA SA."
            sheet.cell(row=current_row, column=24).value = "4050373153151"
            sheet.cell(row=current_row, column=25).value = posic.cultivo
            sheet.cell(row=current_row, column=26).value = conf.codigo_trazabilidad
            
            current_row += 1
            total_pallets_all += 1

    # Guardar en memoria y retornar
    temp_file = io.BytesIO()
    wb.save(temp_file)
    temp_file.seek(0)
    
    clean_nave = re.sub(r'[^a-zA-Z0-9_\-]', '_', nave)
    filename = f"PACKING_LIST_OGL_{clean_nave}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        temp_file, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@router.get("/finalized")
def get_ogl_finalized_naves(db: Session = Depends(get_db)):
    """Lista naves donde todas sus órdenes OGL están finalizadas"""
    ogl_clients = db.query(CatClienteIE).filter(CatClienteIE.nombre_comercial.ilike("%OGL%")).all()
    if not ogl_clients: return []
    ogl_names = [c.nombre_comercial for c in ogl_clients]
    
    # Buscamos todas las naves
    all_naves = db.query(RefPosicionamiento.nave).filter(
        RefPosicionamiento.cliente.in_(ogl_names),
        RefPosicionamiento.nave.isnot(None),
        RefPosicionamiento.nave != ""
    ).distinct().all()
    
    finalized = []
    for nave_row in all_naves:
        nave = nave_row[0]
        # Verificar si TODAS las órdenes de esta nave están finalizadas
        total = db.query(RefPosicionamiento).filter(
            RefPosicionamiento.nave == nave,
            RefPosicionamiento.cliente.in_(ogl_names)
        ).count()
        
        finalizadas = db.query(RefPosicionamiento).filter(
            RefPosicionamiento.nave == nave,
            RefPosicionamiento.cliente.in_(ogl_names),
            RefPosicionamiento.packing_ogl_finalizado == True
        ).count()
        
        if total > 0 and total == finalizadas:
            finalized.append(nave)
            
    return finalized

@router.post("/reopen/{nave}")
def reopen_ogl_nave(nave: str, db: Session = Depends(get_db)):
    """Reabre una nave (anula el bloqueo de finalización) parea permitir corregir archivos"""
    ogl_clients = db.query(CatClienteIE).filter(CatClienteIE.nombre_comercial.ilike("%OGL%")).all()
    ogl_names = [c.nombre_comercial for c in ogl_clients]
    
    rows = db.query(RefPosicionamiento).filter(
        RefPosicionamiento.nave == nave,
        RefPosicionamiento.cliente.in_(ogl_names)
    ).all()
    
    if not rows:
        raise HTTPException(status_code=404, detail="Nave no encontrada")
        
    for r in rows:
        r.packing_ogl_finalizado = False
        
    db.commit()
    return {"ok": True, "message": f"Nave {nave} reabierta con éxito"}
