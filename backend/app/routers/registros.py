from __future__ import annotations

from datetime import datetime, timezone, date, time, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import APIRouter, Depends, HTTPException, Query, Header
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import or_, desc, func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.dependencies.auth import get_current_user_optional
from app.models.auth import Usuario
from app.models.catalogos import Chofer, Vehiculo, Transportista
from app.models.operacion import RegistroOperativo
from app.models.auditoria import RegistroEvento
from app.models.unicos import Unico
from app.models.ref_posicionamiento import RefPosicionamiento
from app.models.ref_booking_dam import RefBookingDam
from app.schemas.operacion import RegistroCrear, RegistroRespuesta, FilaSapRespuesta
from app.schemas.historial import RegistroListado, HistorialResponse
from app.utils.unicidad import normalizar, dividir_por_slash, unir_por_slash

router = APIRouter(prefix="/api/v1/registros", tags=["Registros"])

# Valores “históricos”: se bloquean si se repiten (siempre)
TIPOS_HISTORICOS = {
    "O_BETA",
    "BOOKING",
    "TERMOGRAFO",
    "PS_BETA",
    "PS_ADUANA",
    "PS_OPERADOR",
    "SENASA_PS_LINEA",
}

# Valores “vigentes”: se bloquean solo mientras el registro está “en curso”
# (AWB queda libre al PROCESAR/ANULAR, o automático tras DIAS_TRAVESIA_AWB).
TIPOS_VIGENTES = {"AWB"}

# Días de travesía típica: tras este tiempo el AWB se considera libre aunque no se haya procesado/anulado.
DIAS_TRAVESIA_AWB = 35  # 1 mes y 5 días


# ===============================
# Helpers
# ===============================
def registrar_evento(
    db: Session,
    registro_id: int,
    accion: str,
    antes: dict | None = None,
    despues: dict | None = None,
    motivo: str | None = None,
    usuario: str | None = "sistema",
):
    db.add(
        RegistroEvento(
            registro_id=registro_id,
            accion=accion,
            antes=antes,
            despues=despues,
            motivo=(motivo or None),
            usuario=(usuario or None),
        )
    )


def construir_senasa_ps_linea(senasa: str | None, ps_linea: str | None) -> str | None:
    senasa = (senasa or "").strip()
    ps_linea = (ps_linea or "").strip()
    if senasa and ps_linea:
        return f"{senasa}/PS.{ps_linea}"
    if ps_linea:
        return f"PS.{ps_linea}"
    return None


def safe_str(x) -> str:
    return (x or "").strip()


def obtener_refs_por_booking(db: Session, booking: str | None) -> dict:
    b = normalizar(booking)
    if not b:
        return {"booking": None, "o_beta": None, "awb": None, "dam": None}

    pos = db.query(RefPosicionamiento).filter(RefPosicionamiento.booking == b).first()
    dam = db.query(RefBookingDam).filter(RefBookingDam.booking == b).first()

    return {
        "booking": b,
        "o_beta": normalizar(pos.o_beta) if pos and pos.o_beta else None,
        "awb": normalizar(pos.awb) if pos and pos.awb else None,
        "dam": normalizar(dam.dam) if dam and dam.dam else None,
    }


def construir_items_unicos(payload: RegistroCrear, senasa_ps_linea_norm: str | None) -> list[tuple[str, str, bool]]:
    items: list[tuple[str, str, bool]] = []

    def add(tipo: str, valor: str | None):
        v = normalizar(valor)
        if not v:
            return
        # IGNORAR ASTERISCOS (ej: *, **, ***, ****) en validación de unicidad
        if set(v) == {'*'}:
            return

        vigente = tipo in TIPOS_VIGENTES
        items.append((tipo, v, vigente))

    add("O_BETA", payload.o_beta)
    add("BOOKING", payload.booking)
    add("AWB", payload.awb)

    for t in dividir_por_slash(payload.termografos):
        add("TERMOGRAFO", t)

    for ps in dividir_por_slash(payload.ps_beta):
        add("PS_BETA", ps)

    add("PS_ADUANA", payload.ps_aduana)
    add("PS_OPERADOR", payload.ps_operador)
    add("SENASA_PS_LINEA", senasa_ps_linea_norm)

    return items


def validar_duplicados(db: Session, items: list[tuple[str, str, bool]]) -> list[dict]:
    duplicados: list[dict] = []
    ahora = datetime.now(timezone.utc)
    limite_travesia = ahora - timedelta(days=DIAS_TRAVESIA_AWB)

    for tipo, valor, vigente in items:
        q = db.query(Unico).filter(Unico.tipo == tipo, Unico.valor == valor)
        if vigente:
            # AWB: solo bloquea si vigente Y fecha_uso dentro de los últimos DIAS_TRAVESIA_AWB
            q = q.filter(
                Unico.vigente == True,  # noqa: E712
                Unico.fecha_uso >= limite_travesia,
            )
        existe = q.first()
        if existe:
            duplicados.append(
                {
                    "tipo": tipo,
                    "valor": valor,
                    "mensaje": "Valor en uso actualmente (candado vigente)"
                    if vigente
                    else "Valor ya utilizado (bloqueado por unicidad)",
                }
            )
    return duplicados


def validar_duplicados_excluyendo(
    db: Session, items: list[tuple[str, str, bool]], referencia: str
) -> list[dict]:
    duplicados: list[dict] = []
    ahora = datetime.now(timezone.utc)
    limite_travesia = ahora - timedelta(days=DIAS_TRAVESIA_AWB)

    for tipo, valor, vigente in items:
        q = db.query(Unico).filter(
            Unico.tipo == tipo,
            Unico.valor == valor,
            Unico.referencia != referencia,
        )
        if vigente:
            q = q.filter(
                Unico.vigente == True,  # noqa: E712
                Unico.fecha_uso >= limite_travesia,
            )
        existe = q.first()
        if existe:
            duplicados.append(
                {
                    "tipo": tipo,
                    "valor": valor,
                    "mensaje": "Valor en uso actualmente (candado vigente)"
                    if vigente
                    else "Valor ya utilizado (bloqueado por unicidad)",
                }
            )
    return duplicados


def snapshot_registro(reg: RegistroOperativo) -> dict:
    return {
        "id": reg.id,
        "estado": reg.estado,
        "booking": reg.booking,
        "o_beta": reg.o_beta,
        "awb": reg.awb,
        "dam": reg.dam,
        "chofer_id": reg.chofer_id,
        "transportista_id": reg.transportista_id,
        "termografos": reg.termografos,
        "ps_beta": reg.ps_beta,
        "ps_aduana": reg.ps_aduana,
        "ps_operador": reg.ps_operador,
        "senasa": reg.senasa,
        "ps_linea": reg.ps_linea,
        "senasa_ps_linea": reg.senasa_ps_linea,
        "processed_at": reg.processed_at.isoformat() if getattr(reg, "processed_at", None) else None,
        "anulado_at": reg.anulado_at.isoformat() if getattr(reg, "anulado_at", None) else None,
        "anulado_motivo": reg.anulado_motivo,
    }


def construir_items_unicos_desde_reg(reg: RegistroOperativo) -> list[tuple[str, str, bool]]:
    items: list[tuple[str, str, bool]] = []

    def add(tipo: str, valor: str | None):
        v = normalizar(valor)
        if not v:
            return
        vigente = tipo in TIPOS_VIGENTES
        items.append((tipo, v, vigente))

    add("O_BETA", reg.o_beta)
    add("BOOKING", reg.booking)
    add("AWB", reg.awb)

    for t in dividir_por_slash(reg.termografos):
        add("TERMOGRAFO", t)
    for ps in dividir_por_slash(reg.ps_beta):
        add("PS_BETA", ps)

    add("PS_ADUANA", reg.ps_aduana)
    add("PS_OPERADOR", reg.ps_operador)
    add("SENASA_PS_LINEA", reg.senasa_ps_linea)

    return items


def recrear_unicos_del_registro(db: Session, reg: RegistroOperativo):
    """
    Reconstruye unicidad del registro completo, revalidando colisiones.
    Regla clave: si el registro ya está PROCESADO/ANULADO/CERRADO, entonces
    lo “vigente” (AWB) se marca como no vigente (se libera candado).
    """
    referencia = f"REG-{reg.id}"
    items = construir_items_unicos_desde_reg(reg)

    if reg.estado in {"procesado", "anulado", "cerrado"}:
        items = [(t, v, False if t in TIPOS_VIGENTES else vig) for (t, v, vig) in items]

    duplicados = validar_duplicados_excluyendo(db, items, referencia)
    if duplicados:
        raise HTTPException(status_code=409, detail={"duplicados": duplicados})

    db.query(Unico).filter(Unico.referencia == referencia).delete(synchronize_session=False)

    for tipo, valor, vigente in items:
        db.add(
            Unico(
                tipo=tipo,
                valor=valor,
                referencia=referencia,
                usuario="sistema",
                origen="registro",
                vigente=vigente,
            )
        )


# ===============================
# CRUD / Dominio
# ===============================
@router.post("", response_model=RegistroRespuesta)
def crear_registro(payload: RegistroCrear, db: Session = Depends(get_db)):
    chofer = db.query(Chofer).filter(Chofer.dni == payload.dni).first()
    if not chofer:
        raise HTTPException(status_code=404, detail="Chofer no encontrado por DNI")

    # El transportista se determina por la placa TRACTO; ambas placas son obligatorias para registrar
    placas_raw = (payload.placas or "").strip().upper()
    parts = [p.strip() for p in placas_raw.split("/")] if placas_raw else []
    tracto = parts[0] if len(parts) > 0 else ""
    carreta = parts[1] if len(parts) > 1 else ""
    if not tracto or not carreta:
        raise HTTPException(
            status_code=422,
            detail="Debes enviar placa tracto y placa carreta (ambas obligatorias para registrar)",
        )

    vehiculo = (
        db.query(Vehiculo)
        .options(joinedload(Vehiculo.transportista))
        .filter(Vehiculo.placa_tracto == tracto)
        .first()
    )
    if not vehiculo:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado para esa placa tracto")

    transportista = vehiculo.transportista
    if not transportista:
        raise HTTPException(
            status_code=422,
            detail="El vehículo no tiene transportista asociado. Asocia el transportista al vehículo en catálogos.",
        )

    # Autocompletar desde refs por booking
    refs = obtener_refs_por_booking(db, payload.booking)
    if refs["booking"]:
        payload.booking = refs["booking"]
        if not normalizar(payload.o_beta) and refs["o_beta"]:
            payload.o_beta = refs["o_beta"]
        if not normalizar(payload.awb) and refs["awb"]:
            payload.awb = refs["awb"]
        if hasattr(payload, "dam"):
            if not normalizar(getattr(payload, "dam", None)) and refs["dam"]:
                payload.dam = refs["dam"]

    senasa_ps_linea = construir_senasa_ps_linea(payload.senasa, payload.ps_linea)
    senasa_ps_linea_norm = normalizar(senasa_ps_linea)

    o_beta_norm = normalizar(payload.o_beta)
    booking_norm = normalizar(payload.booking)
    awb_norm = normalizar(payload.awb)

    termografos_norm = unir_por_slash(dividir_por_slash(payload.termografos))
    ps_beta_norm = unir_por_slash(dividir_por_slash(payload.ps_beta))

    ps_aduana_norm = normalizar(payload.ps_aduana)
    ps_operador_norm = normalizar(payload.ps_operador)

    senasa_norm = normalizar(payload.senasa)
    ps_linea_norm = normalizar(payload.ps_linea)

    dam_norm = normalizar(getattr(payload, "dam", None))

    items_unicos = construir_items_unicos(payload, senasa_ps_linea_norm)
    duplicados = validar_duplicados(db, items_unicos)
    if duplicados:
        raise HTTPException(status_code=409, detail={"duplicados": duplicados})

    reg = RegistroOperativo(
        o_beta=o_beta_norm,
        booking=booking_norm,
        awb=awb_norm,
        chofer_id=chofer.id,
        vehiculo_id=vehiculo.id,
        transportista_id=transportista.id,
        termografos=termografos_norm,
        ps_beta=ps_beta_norm,
        ps_aduana=ps_aduana_norm,
        ps_operador=ps_operador_norm,
        senasa=senasa_norm,
        ps_linea=ps_linea_norm,
        senasa_ps_linea=senasa_ps_linea_norm,
        dam=dam_norm,
        estado="pendiente",
    )

    try:
        db.add(reg)
        db.flush()

        referencia = f"REG-{reg.id}"
        for tipo, valor, vigente in items_unicos:
            db.add(
                Unico(
                    tipo=tipo,
                    valor=valor,
                    referencia=referencia,
                    usuario="sistema",
                    origen="registro",
                    vigente=vigente,
                )
            )

        registrar_evento(db, registro_id=reg.id, accion="CREAR", antes=None, despues=snapshot_registro(reg))

        db.commit()
        db.refresh(reg)
        return reg

    except IntegrityError:
        db.rollback()
        duplicados = validar_duplicados(db, items_unicos)
        if duplicados:
            raise HTTPException(status_code=409, detail={"duplicados": duplicados})
        raise HTTPException(status_code=409, detail="Conflicto de unicidad.")


@router.post("/{registro_id}/cerrar")
def cerrar_registro(registro_id: int, db: Session = Depends(get_db)):
    """
    Histórico (compatibilidad): “cerrar” == PROCESAR.
    """
    reg = db.query(RegistroOperativo).filter(RegistroOperativo.id == registro_id).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    if reg.estado in {"procesado", "cerrado"}:
        return {"estado": "ya estaba procesado"}

    if reg.estado == "anulado":
        raise HTTPException(status_code=409, detail="No se puede procesar un registro anulado")

    antes = snapshot_registro(reg)
    reg.estado = "procesado"
    reg.processed_at = datetime.now(timezone.utc)

    # Liberar AWB vigente (candado)
    referencia = f"REG-{reg.id}"
    ahora = datetime.now(timezone.utc)

    db.query(Unico).filter(
        Unico.referencia == referencia,
        Unico.tipo.in_(list(TIPOS_VIGENTES)),
        Unico.vigente == True,  # noqa: E712
    ).update(
        {"vigente": False, "liberado_en": ahora},
        synchronize_session=False,
    )

    registrar_evento(db, registro_id=reg.id, accion="PROCESAR", antes=antes, despues=snapshot_registro(reg))

    db.commit()
    return {"estado": "procesado", "awbs_liberados": True}


@router.post("/{registro_id}/procesar")
def procesar_registro(registro_id: int, db: Session = Depends(get_db)):
    """Endpoint pro (dominio) -> alias del histórico /cerrar."""
    return cerrar_registro(registro_id, db)


class AnularRequest(BaseModel):
    motivo: str


@router.post("/{registro_id}/anular")
def anular_registro(
    registro_id: int,
    payload: AnularRequest,
    db: Session = Depends(get_db),
    current_user: Usuario | None = Depends(get_current_user_optional),
):
    """
    Anular un registro ya PROCESADO (ej: se cargó por error, contenedor no salió, etc.).
    - El registro queda en estado "anulado" (auditoría + motivo).
    - Se libera el AWB (contenedor) para que puedas usarlo de nuevo en otro registro.
    - Se registra quién anuló en ope_registro_eventos (usuario del JWT si hay sesión).
    """
    reg = db.query(RegistroOperativo).filter(RegistroOperativo.id == registro_id).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    if reg.estado != "procesado":
        raise HTTPException(status_code=409, detail="Solo se puede anular un registro en estado PROCESADO")

    motivo = (payload.motivo or "").strip()
    if not motivo:
        raise HTTPException(status_code=422, detail="Debes indicar un motivo de anulación")

    antes = snapshot_registro(reg)
    reg.estado = "anulado"
    reg.anulado_at = datetime.now(timezone.utc)
    reg.anulado_motivo = motivo

    # Liberar AWB (contenedor) para que se pueda reutilizar en otro registro
    referencia = f"REG-{reg.id}"
    ahora = datetime.now(timezone.utc)
    db.query(Unico).filter(
        Unico.referencia == referencia,
        Unico.tipo.in_(list(TIPOS_VIGENTES)),
        Unico.vigente == True,  # noqa: E712
    ).update(
        {"vigente": False, "liberado_en": ahora},
        synchronize_session=False,
    )

    registrar_evento(
        db,
        registro_id=reg.id,
        accion="ANULAR",
        antes=antes,
        despues=snapshot_registro(reg),
        motivo=motivo,
        usuario=current_user.usuario if current_user else "sistema",
    )

    db.commit()
    return {"estado": "anulado", "awbs_liberados": True}


class EditarCampoRequest(BaseModel):
    campo: str
    valor: str


class EditarRegistroRequest(BaseModel):
    """
    Edición controlada (1 “área” por vez), auditable.

    Reglas:
    - Solo registros PROCESADOS pueden editarse.
    - El usuario elige campo/área editable.
    - Al editar BOOKING se recalculan O_BETA/AWB/DAM desde referencias.
    """
    campo: str
    data: dict
    motivo: str | None = None


EDITABLES = {
    "booking",
    "awb",  # contenedor (tu UI lo muestra)
    "dni_chofer",
    "transportista",
    "termografos",
    "precintos",
}


@router.patch("/{registro_id}/editar-campo")
def editar_campo_registro(registro_id: int, payload: EditarCampoRequest, db: Session = Depends(get_db)):
    """
    ENDPOINT LEGADO (compatibilidad).
    Mantenerlo simple y limitado.
    """
    reg = db.query(RegistroOperativo).filter(RegistroOperativo.id == registro_id).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    if reg.estado != "procesado":
        raise HTTPException(status_code=409, detail="Solo se puede editar un registro en estado PROCESADO")

    campo = (payload.campo or "").strip().lower()
    valor = (payload.valor or "").strip()

    if not campo:
        raise HTTPException(status_code=422, detail="Debes indicar el campo a editar")

    allowed = {"booking", "dni", "transportista", "termografos"}
    if campo not in allowed:
        raise HTTPException(status_code=422, detail=f"Campo no editable: {campo}")

    antes = snapshot_registro(reg)

    if campo == "booking":
        refs = obtener_refs_por_booking(db, valor)
        reg.booking = normalizar(valor) if not refs.get("booking") else refs["booking"]
        if refs.get("booking"):
            reg.o_beta = refs.get("o_beta")
            reg.awb = refs.get("awb")
            reg.dam = refs.get("dam")

    elif campo == "dni":
        dni = normalizar(valor)
        chofer = db.query(Chofer).filter(Chofer.dni == dni).first()
        if not chofer:
            raise HTTPException(status_code=404, detail="Chofer no encontrado por DNI")
        reg.chofer_id = chofer.id

    elif campo == "transportista":
        needle = valor.strip()
        if not needle:
            raise HTTPException(status_code=422, detail="Debes indicar el transportista")

        q = db.query(Transportista).filter(
            or_(
                Transportista.ruc == needle,
                Transportista.codigo_sap == needle,
                Transportista.nombre_transportista.ilike(f"%{needle}%"),
            )
        )
        matches = q.limit(6).all()
        if len(matches) == 0:
            raise HTTPException(status_code=404, detail="Transportista no encontrado")
        if len(matches) > 1:
            raise HTTPException(status_code=409, detail="Transportista ambiguo. Usa RUC o Código SAP.")
        reg.transportista_id = matches[0].id

    elif campo == "termografos":
        reg.termografos = unir_por_slash(dividir_por_slash(valor))

    recrear_unicos_del_registro(db, reg)

    registrar_evento(
        db,
        registro_id=reg.id,
        accion="EDITAR",
        antes=antes,
        despues=snapshot_registro(reg),
        usuario=current_user.usuario if current_user else "sistema",
    )

    db.commit()
    return {"ok": True, "registro_id": reg.id, "campo": campo}


# Roles que pueden editar registros procesados (JWT o header X-User-Role).
ROLES_EDITOR = {"admin", "editor"}


def require_rol_editor(
    optional_user: Usuario | None = Depends(get_current_user_optional),
    x_user_role: str | None = Header(default=None, alias="X-User-Role"),
) -> str:
    """Solo usuarios con rol admin o editor pueden editar. Prioridad: JWT > X-User-Role."""
    if optional_user and optional_user.rol in ("administrador", "supervisor_facturacion"):
        return "admin" if optional_user.rol == "administrador" else "editor"
    role = (x_user_role or "").strip().lower()
    if role not in ROLES_EDITOR:
        raise HTTPException(
            status_code=403,
            detail="Sin permiso para editar. Se requiere rol admin o editor.",
        )
    return role


@router.patch("/{registro_id}/editar")
def editar_registro_controlado(
    registro_id: int,
    payload: EditarRegistroRequest,
    db: Session = Depends(get_db),
    _role: str = Depends(require_rol_editor),
    current_user: Usuario | None = Depends(get_current_user_optional),
):
    """
    Edición controlada y auditable (solo roles admin/editor).

    - Requiere header X-User-Role: admin o editor (o JWT con rol administrador/supervisor).
    - Solo estado PROCESADO.
    - Se registra quién editó en ope_registro_eventos (usuario del JWT si hay sesión).
    """
    reg = db.query(RegistroOperativo).filter(RegistroOperativo.id == registro_id).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    if reg.estado != "procesado":
        raise HTTPException(status_code=409, detail="Solo se puede editar un registro en estado PROCESADO")

    campo = (payload.campo or "").strip().lower()
    if campo not in EDITABLES:
        raise HTTPException(status_code=422, detail=f"Campo no editable: {campo}")

    data = payload.data or {}
    antes = snapshot_registro(reg)

    # 1) BOOKING (recalcula refs)
    if campo == "booking":
        booking_raw = (data.get("booking") or "").strip()
        if not booking_raw:
            raise HTTPException(status_code=422, detail="Debes enviar booking")

        refs = obtener_refs_por_booking(db, booking_raw)
        reg.booking = refs.get("booking") or normalizar(booking_raw)
        reg.o_beta = refs.get("o_beta")
        reg.awb = refs.get("awb")
        reg.dam = refs.get("dam")

    # 2) AWB (contenedor)
    elif campo == "awb":
        awb_raw = (data.get("awb") or "").strip()
        reg.awb = normalizar(awb_raw) or None

    # 3) DNI CHOFER
    elif campo == "dni_chofer":
        dni = normalizar((data.get("dni") or "").strip())
        if not dni:
            raise HTTPException(status_code=422, detail="Debes enviar dni")
        chofer = db.query(Chofer).filter(Chofer.dni == dni).first()
        if not chofer:
            raise HTTPException(status_code=404, detail="Chofer no encontrado por DNI")
        reg.chofer_id = chofer.id

    # 4) TRANSPORTISTA
    elif campo == "transportista":
        needle = (data.get("needle") or "").strip()
        if not needle:
            raise HTTPException(status_code=422, detail="Debes enviar needle (RUC o Código SAP)")

        q = db.query(Transportista).filter(
            or_(
                Transportista.ruc == needle,
                Transportista.codigo_sap == needle,
                Transportista.nombre_transportista.ilike(f"%{needle}%"),
            )
        )
        matches = q.limit(6).all()
        if len(matches) == 0:
            raise HTTPException(status_code=404, detail="Transportista no encontrado")
        if len(matches) > 1:
            raise HTTPException(status_code=409, detail="Transportista ambiguo. Usa RUC o Código SAP.")
        reg.transportista_id = matches[0].id

    # 5) TERMÓGRAFOS
    elif campo == "termografos":
        reg.termografos = unir_por_slash(dividir_por_slash(data.get("termografos")))

    # 6) PRECINTOS (agrupado)
    elif campo == "precintos":
        if "ps_beta" in data:
            reg.ps_beta = unir_por_slash(dividir_por_slash(data.get("ps_beta")))

        if "ps_aduana" in data:
            reg.ps_aduana = normalizar(data.get("ps_aduana"))
        if "ps_operador" in data:
            reg.ps_operador = normalizar(data.get("ps_operador"))

        if "senasa" in data:
            reg.senasa = normalizar(data.get("senasa"))
        if "ps_linea" in data:
            reg.ps_linea = normalizar(data.get("ps_linea"))

        reg.senasa_ps_linea = construir_senasa_ps_linea(reg.senasa, reg.ps_linea)

    else:
        raise HTTPException(status_code=422, detail=f"Campo no editable: {campo}")

    recrear_unicos_del_registro(db, reg)

    registrar_evento(
        db,
        registro_id=reg.id,
        accion="EDITAR",
        antes=antes,
        despues=snapshot_registro(reg),
        motivo=(payload.motivo or None),
    )

    db.commit()
    return {"ok": True, "registro_id": reg.id, "campo": campo, "estado": reg.estado}


@router.get("/{registro_id}/sap", response_model=FilaSapRespuesta)
def obtener_fila_sap(registro_id: int, db: Session = Depends(get_db)):
    reg = (
        db.query(RegistroOperativo)
        .options(
            joinedload(RegistroOperativo.chofer),
            joinedload(RegistroOperativo.vehiculo),
            joinedload(RegistroOperativo.transportista),
        )
        .filter(RegistroOperativo.id == registro_id)
        .first()
    )
    if not reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    fecha_txt = reg.fecha_registro.date().isoformat()

    chofer = reg.chofer
    veh = reg.vehiculo
    tra = reg.transportista

    return FilaSapRespuesta(
        FECHA=fecha_txt,
        O_BETA=safe_str(reg.o_beta),
        BOOKING=safe_str(reg.booking),
        AWB=safe_str(reg.awb),
        MARCA=safe_str(veh.marca),
        PLACAS=safe_str(veh.placas),
        DNI=safe_str(chofer.dni),
        CHOFER=chofer.nombre_para_sap,
        LICENCIA=safe_str(chofer.licencia),
        TERMOGRAFOS=safe_str(reg.termografos),
        CODIGO_SAP=safe_str(tra.codigo_sap),
        TRANSPORTISTA=safe_str(tra.nombre_transportista),
        PS_BETA=safe_str(reg.ps_beta),
        PS_ADUANA=safe_str(reg.ps_aduana),
        PS_OPERADOR=safe_str(reg.ps_operador),
        SENASA_PS_LINEA=safe_str(reg.senasa_ps_linea),
        N_DAM=safe_str(reg.dam),
        P_REGISTRAL=safe_str(tra.partida_registral),
        CER_VEHICULAR=safe_str(veh.cert_vehicular),
    )


# -------------------------------------------------------------------
# LISTADO DE PROCESADOS POR FECHA (para BandejaSAP > Procesados)
# -------------------------------------------------------------------
class ProcesadoSapItem(FilaSapRespuesta):
    registro_id: int
    estado: str | None = None
    processed_at: datetime | None = None


class ProcesadosSapResponse(BaseModel):
    items: list[ProcesadoSapItem]
    total: int


@router.get("/procesados", response_model=ProcesadosSapResponse)
def listar_procesados(
    fecha: date = Query(..., description="Fecha local Lima YYYY-MM-DD"),
    limit: int = Query(default=200, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
):
    """
    Devuelve registros con processed_at (incluye PROCESADO y ANULADO) para una fecha dada.

    Nota sobre zona horaria:
    - En entornos con base de datos de zonas (Linux, contenedores, etc.) se usa "America/Lima".
    - En algunos entornos Windows sin tzdata instalado, ZoneInfo puede fallar.
      En ese caso hacemos fallback a UTC para evitar errores 500.
    """
    try:
        lima = ZoneInfo("America/Lima")
    except ZoneInfoNotFoundError:
        lima = timezone.utc

    # Rango del día en la zona elegida (inclusive)
    start_local = datetime.combine(fecha, time.min).replace(tzinfo=lima)
    end_local = datetime.combine(fecha, time.max).replace(tzinfo=lima)

    # A UTC para comparar con processed_at (que se guarda en UTC)
    start_utc = start_local.astimezone(timezone.utc)
    end_utc = end_local.astimezone(timezone.utc)

    # Query base para filtros (sin order_by para el count)
    base_filters = (
        db.query(RegistroOperativo)
        .filter(RegistroOperativo.processed_at.isnot(None))
        .filter(RegistroOperativo.processed_at >= start_utc)
        .filter(RegistroOperativo.processed_at <= end_utc)
    )

    # Count sin ORDER BY (PostgreSQL requiere esto)
    total = base_filters.with_entities(func.count(RegistroOperativo.id)).scalar() or 0

    # Query completa con joins, order_by, paginación
    base = (
        base_filters
        .options(
            joinedload(RegistroOperativo.chofer),
            joinedload(RegistroOperativo.vehiculo),
            joinedload(RegistroOperativo.transportista),
        )
        .order_by(desc(RegistroOperativo.processed_at), desc(RegistroOperativo.id))
    )
    rows = base.offset(offset).limit(limit).all()

    items: list[ProcesadoSapItem] = []
    for r in rows:
        chofer = r.chofer
        veh = r.vehiculo
        tra = r.transportista
        fecha_txt = r.fecha_registro.date().isoformat()

        items.append(
            ProcesadoSapItem(
                registro_id=r.id,
                estado=r.estado,
                processed_at=r.processed_at,
                FECHA=fecha_txt,
                O_BETA=safe_str(r.o_beta),
                BOOKING=safe_str(r.booking),
                AWB=safe_str(r.awb),
                MARCA=safe_str(veh.marca if veh else None),
                PLACAS=safe_str(veh.placas if veh else None),
                DNI=safe_str(chofer.dni if chofer else None),
                CHOFER=(chofer.nombre_para_sap if chofer else "") or "",
                LICENCIA=safe_str(chofer.licencia if chofer else None),
                TERMOGRAFOS=safe_str(r.termografos),
                CODIGO_SAP=safe_str(tra.codigo_sap if tra else None),
                TRANSPORTISTA=safe_str(tra.nombre_transportista if tra else None),
                PS_BETA=safe_str(r.ps_beta),
                PS_ADUANA=safe_str(r.ps_aduana),
                PS_OPERADOR=safe_str(r.ps_operador),
                SENASA_PS_LINEA=safe_str(r.senasa_ps_linea),
                N_DAM=safe_str(r.dam),
                P_REGISTRAL=safe_str(tra.partida_registral if tra else None),
                CER_VEHICULAR=safe_str(veh.cert_vehicular if veh else None),
            )
        )

    return ProcesadosSapResponse(items=items, total=total)


@router.get("/historial", response_model=HistorialResponse)
def listar_historial(
    desde: date | None = Query(default=None),
    hasta: date | None = Query(default=None),
    limit: int = Query(default=10, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
):
    """Listado paginado para dashboard (10/25/50/100 por página)."""
    base = db.query(RegistroOperativo)
    if desde:
        base = base.filter(RegistroOperativo.fecha_registro >= datetime.combine(desde, time.min))
    if hasta:
        base = base.filter(RegistroOperativo.fecha_registro <= datetime.combine(hasta, time.max))

    total = base.with_entities(func.count(RegistroOperativo.id)).scalar() or 0

    q = (
        base.options(
            joinedload(RegistroOperativo.chofer),
            joinedload(RegistroOperativo.vehiculo),
            joinedload(RegistroOperativo.transportista),
        )
        .order_by(desc(RegistroOperativo.fecha_registro), desc(RegistroOperativo.id))
    )
    rows = q.offset(offset).limit(limit).all()

    out: list[RegistroListado] = []
    for r in rows:
        out.append(
            RegistroListado(
                id=r.id,
                fecha_registro=r.fecha_registro,
                estado=r.estado,
                booking=r.booking,
                o_beta=r.o_beta,
                awb=r.awb,
                dam=r.dam,
                dni=r.chofer.dni if r.chofer else None,
                chofer=r.chofer.nombre_para_sap if r.chofer else None,
                placas=r.vehiculo.placas if r.vehiculo else None,
                transportista=r.transportista.nombre_transportista if r.transportista else None,
                codigo_sap=r.transportista.codigo_sap if r.transportista else None,
            )
        )

    return HistorialResponse(items=out, total=total)


# ===============================
# Dashboard estadísticas (gráficos por rol)
# ===============================
class DashboardStatsPorDia(BaseModel):
    fecha: str  # YYYY-MM-DD
    total: int
    pendientes: int
    procesados: int
    anulados: int


class DashboardStatsPorEstado(BaseModel):
    estado: str
    total: int


class DashboardStatsPorTransportista(BaseModel):
    nombre: str
    total: int


class DashboardStatsResponse(BaseModel):
    por_dia: list[DashboardStatsPorDia]
    por_estado: list[DashboardStatsPorEstado]
    por_transportista: list[DashboardStatsPorTransportista]
    total_registros: int


@router.get("/dashboard-stats", response_model=DashboardStatsResponse)
def dashboard_stats(
    desde: date | None = Query(default=None),
    hasta: date | None = Query(default=None),
    dias: int = Query(default=30, ge=7, le=90),
    db: Session = Depends(get_db),
):
    """
    Estadísticas agregadas para el dashboard (gráficos por día, estado, transportista).
    Si no se envían desde/hasta, se usan los últimos `dias` días.
    """
    try:
        tz = ZoneInfo("America/Lima")
    except ZoneInfoNotFoundError:
        tz = timezone.utc
    now = datetime.now(tz)
    if hasta:
        end_date = hasta
    else:
        end_date = now.date()
    if desde:
        start_date = desde
    else:
        start_date = end_date - timedelta(days=dias)

    start_dt = datetime.combine(start_date, time.min).replace(tzinfo=tz)
    end_dt = datetime.combine(end_date, time.max).replace(tzinfo=tz)

    q = (
        db.query(RegistroOperativo)
        .options(joinedload(RegistroOperativo.transportista))
        .filter(
            RegistroOperativo.fecha_registro >= start_dt,
            RegistroOperativo.fecha_registro <= end_dt,
        )
    )
    rows = q.all()

    # Agregar por día (fecha en timezone del servidor)
    por_dia_map: dict[str, dict[str, int]] = {}
    current = start_date
    while current <= end_date:
        key = current.isoformat()
        por_dia_map[key] = {"fecha": key, "total": 0, "pendientes": 0, "procesados": 0, "anulados": 0}
        current = current + timedelta(days=1)

    por_estado: dict[str, int] = {}
    por_transportista: dict[str, int] = {}

    for r in rows:
        # Fecha del registro (en tz)
        fd = r.fecha_registro.astimezone(tz).date() if r.fecha_registro else None
        if fd:
            key = fd.isoformat()
            if key in por_dia_map:
                por_dia_map[key]["total"] += 1
                e = (r.estado or "pendiente").lower()
                if e == "pendiente":
                    por_dia_map[key]["pendientes"] += 1
                elif e == "procesado":
                    por_dia_map[key]["procesados"] += 1
                elif e == "anulado":
                    por_dia_map[key]["anulados"] += 1

        e = (r.estado or "pendiente").strip().lower()
        por_estado[e] = por_estado.get(e, 0) + 1

        name = (r.transportista.nombre_transportista if r.transportista else "Sin asignar").strip() or "Sin asignar"
        por_transportista[name] = por_transportista.get(name, 0) + 1

    por_dia_list = [
        DashboardStatsPorDia(
            fecha=data["fecha"],
            total=data["total"],
            pendientes=data["pendientes"],
            procesados=data["procesados"],
            anulados=data["anulados"],
        )
        for data in sorted(por_dia_map.values(), key=lambda x: x["fecha"])
    ]
    por_estado_list = [DashboardStatsPorEstado(estado=k, total=v) for k, v in sorted(por_estado.items())]
    # Top 10 transportistas
    por_transportista_sorted = sorted(por_transportista.items(), key=lambda x: -x[1])[:10]
    por_transportista_list = [DashboardStatsPorTransportista(nombre=k, total=v) for k, v in por_transportista_sorted]

    return DashboardStatsResponse(
        por_dia=por_dia_list,
        por_estado=por_estado_list,
        por_transportista=por_transportista_list,
        total_registros=len(rows),
    )


# ===============================
# Export XLSX SAP
# ===============================
class ExportSapRequest(BaseModel):
    registro_ids: list[int]


@router.post("/export/sap-xlsx")
def exportar_sap_xlsx(payload: ExportSapRequest, db: Session = Depends(get_db)):
    ids = [int(x) for x in payload.registro_ids if isinstance(x, int) or str(x).isdigit()]
    if not ids:
        raise HTTPException(status_code=422, detail="No se enviaron registro_ids válidos")

    regs = (
        db.query(RegistroOperativo)
        .options(
            joinedload(RegistroOperativo.chofer),
            joinedload(RegistroOperativo.vehiculo),
            joinedload(RegistroOperativo.transportista),
        )
        .filter(RegistroOperativo.id.in_(ids))
        .order_by(desc(RegistroOperativo.fecha_registro), desc(RegistroOperativo.id))
        .all()
    )

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "SAP"

    headers = [
        "FECHA", "O_BETA", "BOOKING", "AWB", "MARCA", "PLACAS", "DNI", "CHOFER", "LICENCIA",
        "TERMOGRAFOS", "CODIGO_SAP", "TRANSPORTISTA", "PS_BETA", "PS_ADUANA", "PS_OPERADOR",
        "SENASA_PS_LINEA", "N_DAM", "P_REGISTRAL", "CER_VEHICULAR"
    ]
    ws.append(headers)

    bold = Font(bold=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in regs:
        chofer = r.chofer
        veh = r.vehiculo
        tra = r.transportista
        fecha_txt = r.fecha_registro.date().isoformat()

        ws.append([
            fecha_txt,
            safe_str(r.o_beta),
            safe_str(r.booking),
            safe_str(r.awb),
            safe_str(veh.marca if veh else None),
            safe_str(veh.placas if veh else None),
            safe_str(chofer.dni if chofer else None),
            (chofer.nombre_para_sap if chofer else "") or "",
            safe_str(chofer.licencia if chofer else None),
            safe_str(r.termografos),
            safe_str(tra.codigo_sap if tra else None),
            safe_str(tra.nombre_transportista if tra else None),
            safe_str(r.ps_beta),
            safe_str(r.ps_aduana),
            safe_str(r.ps_operador),
            safe_str(r.senasa_ps_linea),
            safe_str(r.dam),
            safe_str(tra.partida_registral if tra else None),
            safe_str(veh.cert_vehicular if veh else None),
        ])

    for i, _h in enumerate(headers, start=1):
        max_len = max(
            len(str(ws.cell(row=row, column=i).value or ""))
            for row in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max(12, max_len + 2), 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"sap_export_{datetime.now().date().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
