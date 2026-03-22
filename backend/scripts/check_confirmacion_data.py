import openpyxl
import os

file_path = r"d:\PROJECTS\BETA\BETA LogiCapture 1.0\backend\assets\templates\CONFIRMACION.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb["FORMATO"]
header_row = 6

for r in range(header_row + 1, header_row + 10):
    beta = sheet.cell(row=r, column=5).value
    hu = sheet.cell(row=r, column=6).value
    if hu:
        print(f"R{r}: Beta={beta}, HU={hu}")
