import openpyxl
import os

file_path = r"d:\PROJECTS\BETA\BETA LogiCapture 1.0\backend\assets\templates\TERMOGRAFOS.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active
header_row = 2

for r in range(header_row + 1, header_row + 20):
    beta = sheet.cell(row=r, column=6).value
    term = sheet.cell(row=r, column=13).value
    pall = sheet.cell(row=r, column=14).value
    if beta:
        print(f"R{r}: Orden={beta}, Term={term}, Pallet={pall}")
