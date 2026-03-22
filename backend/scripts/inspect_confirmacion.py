import openpyxl
import os

file_path = r"d:\PROJECTS\BETA\BETA LogiCapture 1.0\backend\assets\templates\CONFIRMACION.xlsx"

if not os.path.exists(file_path):
    print(f"Error: No existe el archivo en {file_path}")
else:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    print(f"Hoja: {sheet.title}")
    
    # Leemos cabecera (primera fila con datos significativos)
    # Buscamos la fila donde estén los nombres de columnas
    header_row = 1
    found = False
    for r in range(1, 10):
        for c in range(1, 20):
            val = sheet.cell(row=r, column=c).value
            if val and ("PALLET" in str(val).upper() or "KILOS" in str(val).upper()):
                header_row = r
                found = True
                break
        if found: break
    
    print(f"Fila de Cabecera detectada: {header_row}")
    for c in range(1, 30):
        val = sheet.cell(row=header_row, column=c).value
        if val:
            print(f"Col {c}: {val}")
    
    # Mostramos las primeras 2 filas de datos
    print("\n--- Primeras 2 filas de datos ---")
    for r in range(header_row + 1, header_row + 3):
        row_data = []
        for c in range(1, 30):
            row_data.append(str(sheet.cell(row=r, column=c).value))
        print(f"R{r}: {row_data}")
