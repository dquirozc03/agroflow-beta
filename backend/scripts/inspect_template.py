import openpyxl
import os

template_path = r"d:\PROJECTS\BETA\BETA LogiCapture 1.0\backend\assets\templates\FORMATO PL - OGL.xlsx"

if not os.path.exists(template_path):
    print(f"Error: No existe el archivo en {template_path}")
else:
    wb = openpyxl.load_all_data(template_path) if hasattr(openpyxl, 'load_all_data') else openpyxl.load_workbook(template_path, data_only=True)
    sheet = wb.active
    print(f"Inspeccionando hoja: {sheet.title}")
    
    # Listamos celdas que parecen tener etiquetas (etiquetas de campos)
    for r in range(1, 40):
        row_values = []
        for c in range(1, 15):
            val = sheet.cell(row=r, column=c).value
            if val:
                print(f"[{openpyxl.utils.cell.get_column_letter(c)}{r}]: {val}")
